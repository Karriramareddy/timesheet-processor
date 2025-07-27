from flask import Flask, request, jsonify
import json
from datetime import datetime
import os
from werkzeug.utils import secure_filename
import tempfile
from openpyxl import load_workbook
import xlrd

app = Flask(__name__)

# Configure upload settings
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_timesheet_data(file_path):
    """
    Parse the Excel timesheet and extract data in the required format
    """
    try:
        timesheet_data = []
        
        # Determine file type and read accordingly
        if file_path.endswith('.xlsx'):
            # Use openpyxl for .xlsx files
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Convert to list of lists for consistent processing
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row) if row else [])
        
        else:
            # Use xlrd for .xls files
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            
            # Convert to list of lists
            data = []
            for row_idx in range(sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    try:
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row.append(cell_value)
                    except:
                        row.append("")
                data.append(row)
        
        # Extract header information
        employee_name = ""
        sesa_id = ""
        reporting_manager_schneider = ""
        project_name = ""
        wissen_employee_id = ""
        reporting_manager_wissen = ""
        
        # Parse header section (first few rows)
        for i in range(min(10, len(data))):
            for j in range(min(10, len(data[i]) if data[i] else 0)):
                try:
                    cell_value = str(data[i][j]).strip() if data[i][j] is not None else ""
                    
                    # Extract employee information based on position
                    if i == 1 and j == 1:  # Employee name position
                        employee_name = cell_value
                    elif i == 2 and j == 1:  # SESA ID position  
                        sesa_id = cell_value
                    elif "Employee Name" in cell_value and j + 1 < len(data[i]):
                        employee_name = str(data[i][j + 1]) if data[i][j + 1] else ""
                    elif "SESA" in cell_value and j + 1 < len(data[i]):
                        sesa_id = str(data[i][j + 1]) if data[i][j + 1] else ""
                    elif "Project" in cell_value and j + 1 < len(data[i]):
                        project_name = str(data[i][j + 1]) if data[i][j + 1] else ""
                except:
                    continue
        
        # Find the timesheet data section
        date_row_index = -1
        
        # Look for the row containing month headers
        for i in range(len(data)):
            if not data[i]:
                continue
            row_has_dates = False
            for j in range(len(data[i])):
                try:
                    cell_value = str(data[i][j]).strip() if data[i][j] is not None else ""
                    # Look for month patterns
                    if any(month in cell_value for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                                            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']):
                        row_has_dates = True
                        break
                except:
                    continue
            
            if row_has_dates:
                date_row_index = i
                break
        
        if date_row_index == -1:
            date_row_index = 5  # Default fallback
        
        # Extract date headers
        date_headers = []
        month_year = ""
        
        if date_row_index < len(data):
            for j in range(len(data[date_row_index])):
                try:
                    cell_value = str(data[date_row_index][j]).strip() if data[date_row_index][j] is not None else ""
                    if cell_value and cell_value != "None":
                        date_headers.append((j, cell_value))
                        if any(month in cell_value for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                                                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']):
                            month_year = cell_value
                except:
                    continue
        
        # Extract timesheet values
        for i in range(date_row_index + 1, len(data)):
            if not data[i]:
                continue
                
            try:
                day_number = str(data[i][0]).strip() if data[i][0] is not None else ""
                
                if not day_number or day_number == "None":
                    continue
                
                # Try to convert to day number
                try:
                    day_num = int(float(day_number))
                except (ValueError, TypeError):
                    continue
                
                # Extract timesheet values for this day
                for col_index, date_header in date_headers:
                    if col_index < len(data[i]):
                        try:
                            timesheet_value = data[i][col_index]
                            
                            # Skip empty cells
                            if timesheet_value is None or str(timesheet_value).strip() == "" or str(timesheet_value) == "None":
                                continue
                            
                            # Convert timesheet value
                            if isinstance(timesheet_value, (int, float)):
                                ts_value = float(timesheet_value)
                            else:
                                ts_str = str(timesheet_value).strip()
                                if ts_str and ts_str != "None":
                                    try:
                                        ts_value = float(ts_str)
                                    except ValueError:
                                        ts_value = ts_str
                                else:
                                    continue
                            
                            # Create the output record
                            record = {
                                "employee_name": employee_name,
                                "SESA_ID": sesa_id,
                                "reporting_manager-schneider": reporting_manager_schneider,
                                "project_name": project_name,
                                "wissen_employee_id": wissen_employee_id,
                                "reporting_manager-wissen": reporting_manager_wissen,
                                "Month": month_year,
                                "Date": f"{day_num:02d}",
                                "timesheet_value": ts_value
                            }
                            
                            timesheet_data.append(record)
                        except:
                            continue
            except:
                continue
        
        return timesheet_data
    
    except Exception as e:
        print(f"Error parsing timesheet: {str(e)}")
        return []

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        "message": "Timesheet Processor API",
        "endpoints": {
            "/upload": "POST - Upload Excel timesheet file",
            "/health": "GET - Health check"
        }
    })

@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})

@app.route('/upload', methods=['POST'])
def upload_timesheet():
    try:
        # Check if file is present in request
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        if not allowed_file(file.filename):
            return jsonify({"error": "Invalid file type. Only .xls and .xlsx files are allowed"}), 400
        
        # Create a temporary file to save the uploaded file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx' if file.filename.endswith('.xlsx') else '.xls') as temp_file:
            file.save(temp_file.name)
            temp_file_path = temp_file.name
        
        try:
            # Process the timesheet
            timesheet_data = parse_timesheet_data(temp_file_path)
            
            if not timesheet_data:
                return jsonify({"error": "No valid timesheet data found in the file"}), 400
            
            # Return the processed data as JSON
            response_data = {
                "status": "success",
                "total_records": len(timesheet_data),
                "data": timesheet_data,
                "processed_at": datetime.now().isoformat()
            }
            
            return jsonify(response_data)
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
    
    except Exception as e:
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
